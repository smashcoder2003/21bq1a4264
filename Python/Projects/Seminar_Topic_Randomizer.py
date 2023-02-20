import random
import openpyxl


def get_seminar_topics(File_location=None, Save_File_Name=None):

    def Topiclist(File_Name=File_location):
        wb = openpyxl.load_workbook(File_Name)
        sh = wb["Sheet1"]
        topics = []
        for col in sh['A']:
            topics.append(col.value)
        return topics

    def create_roll_no_list(File_Name=File_location):
        wb = openpyxl.load_workbook(File_Name)
        sh = wb["Sheet1"]
        roll_nos = []
        for col in sh['B']:
            roll_nos.append(col.value)
        return roll_nos

    def Assign_Topics():
        topics = Topiclist("/Applications/seminar2.xlsx")
        topics.pop(0)
        roll_no_list = create_roll_no_list()
        Topic_Sheet = openpyxl.Workbook()
        sheet = Topic_Sheet.active
        sheet.append(['SNO', 'TopicName', 'RollNo'])
        for x in enumerate(roll_no_list):
            rnd = random.choice(topics)
            sheet.append([x[0]+1, rnd, x[1]])
            topics.remove(rnd)

        sheet.save(Save_File_Name)
    Assign_Topics()


if __name__ == '__main__':
    # get_seminar_topics takes two parameters.
    # Excel file name in which topics, roll_nos are stored.
    # Name of the file in which assigned seminar topics should be stored.
    get_seminar_topics("/Applications/seminar2.xlsx", "AISeminarTopics.xlsx")
